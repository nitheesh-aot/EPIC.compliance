"""Shared utility functions for report generation."""
from flask import current_app

from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.utils.cell import coordinate_from_string

from compliance_api.services.epic_track_service.track_service import TrackService


def sort_dataframe_for_pivot_order(data_frame, column_orders):
    """Sort DataFrame rows to control pivot table column field order.

    Excel builds pivot cache item order from first-occurrence in source data.
    By sorting the dataframe so desired values appear first, we ensure the
    pivot cache (and thus column display) follows the desired order.

    Args:
        data_frame: pandas DataFrame to sort
        column_orders: dict mapping column names to desired value order lists
            e.g. {"compliance_finding": ["In", "Out", "Not Determined"]}

    Returns:
        Sorted DataFrame (original is not modified)
    """
    if data_frame.empty:
        return data_frame

    df = data_frame.copy()

    # Build a combined sort key
    # For each column in column_orders, assign a sort priority based on position
    for col, desired_order in column_orders.items():
        if col not in df.columns:
            continue

        # Create mapping: value -> sort priority (lower = first)
        order_map = {val: idx for idx, val in enumerate(desired_order)}
        # Values not in desired_order get a high priority (appear after)
        max_priority = len(desired_order)

        # Add temporary sort column
        sort_col = f"_sort_{col}"
        df[sort_col] = df[col].map(lambda v: order_map.get(v, max_priority))

    # Sort by all the temporary columns
    sort_cols = [c for c in df.columns if c.startswith("_sort_")]
    if sort_cols:
        df = df.sort_values(by=sort_cols, kind="stable")
        # Remove temporary columns
        df = df.drop(columns=sort_cols)

    return df.reset_index(drop=True)


def get_project_details(project_map, row):
    """Resolve project name/type from approved and unapproved project sources."""
    # Project Logic:
    # If case_file.project_id is null it is unapproved
    # and you should be able to find it in the unapproved_projects table
    # If case_file.project_id has a valid number it should be from EPIC.Track
    if not row.project_id:
        project_name = row.unapproved_project_name
        project_type = row.unapproved_project_type
    else:
        if row.project_id in project_map:
            project = project_map[row.project_id]
        else:
            date = row.case_file_date_created.date() if row.case_file_date_created else None
            project = TrackService.get_project_by_id(row.project_id, as_of_date=date)
            project_map[row.project_id] = project
        project_name = project.get("name") if project else None
        project_type = project.get("type", None).get("name", None) if project else None

    return project_name, project_type


@staticmethod
def populate_template_table_sheet(workbook, sheet_name, data_frame, columns, headers):
    """Populate a template worksheet table while preserving workbook pivot structure."""
    worksheet = workbook[sheet_name]
    table = next(iter(worksheet.tables.values()), None)

    if table is None:
        raise ValueError(f"Template sheet '{sheet_name}' does not contain an Excel table.")

    min_col = 1
    min_row = 1
    table_column_count = len(getattr(table, "tableColumns", []) or [])
    if table_column_count < 1:
        raise ValueError(f"Template sheet '{sheet_name}' table has no columns.")
    max_col = table_column_count

    table_headers = [
        worksheet.cell(row=min_row, column=column_number).value
        for column_number in range(min_col, max_col + 1)
    ]

    header_to_column_key = dict(zip(headers, columns))

    special_headers = {"Index", "Unique Key"}
    unknown_headers = [
        header for header in table_headers
        if header not in special_headers and header not in header_to_column_key
    ]
    if unknown_headers:
        raise ValueError(
            f"Template sheet '{sheet_name}' has unmapped headers: {unknown_headers}"
        )

    records = data_frame.reindex(columns=columns).to_dict("records")

    last_used_row = max(worksheet.max_row, min_row + 1)
    for row_number in range(min_row + 1, last_used_row + 1):
        for column_number in range(min_col, max_col + 1):
            worksheet.cell(row=row_number, column=column_number).value = None

    for row_index, record in enumerate(records, start=1):
        excel_row = min_row + row_index
        for col_offset, header in enumerate(table_headers):
            column_number = min_col + col_offset
            if header == "Index":
                value = row_index
            elif header == "Unique Key":
                value = record.get("enforcement_document_number")
            else:
                column_key = header_to_column_key[header]
                value = record.get(column_key)
            worksheet.cell(row=excel_row, column=column_number).value = value

    new_last_row = min_row + max(len(records), 1)
    table.ref = (
        f"{get_column_letter(min_col)}{min_row}:"
        f"{get_column_letter(max_col)}{new_last_row}"
    )


def reorder_pivot_column_items(workbook, field_name, desired_order):
    """Reorder pivot field items to enforce a specific column display order.

    Sets sortType='manual' and reorders items in pivotField.items to control
    the display order. Does NOT set refreshOnLoad to avoid Excel rebuilding
    the item order from cache.
    """
    # Collect all unique pivot caches from worksheets
    pivot_cache_pairs = []  # List of (cache, pivot) tuples

    for worksheet in workbook.worksheets:
        if not hasattr(worksheet, '_pivots'):
            continue
        for pivot in worksheet._pivots:
            cache = pivot.cache
            if cache is not None:
                pivot_cache_pairs.append((cache, pivot))

    # Group pivots by cache
    cache_to_pivots = {}
    for cache, pivot in pivot_cache_pairs:
        cache_id = id(cache)
        if cache_id not in cache_to_pivots:
            cache_to_pivots[cache_id] = {'cache': cache, 'pivots': []}
        cache_to_pivots[cache_id]['pivots'].append(pivot)

    for cache_data in cache_to_pivots.values():
        cache = cache_data['cache']
        pivots = cache_data['pivots']

        # Find the cache field matching field_name and record its index
        field_idx = None
        cache_field = None
        for idx, cf in enumerate(cache.cacheFields):
            if cf.name == field_name:
                field_idx = idx
                cache_field = cf
                break

        if field_idx is None or cache_field is None:
            continue
        if not hasattr(cache_field, 'sharedItems') or cache_field.sharedItems is None:
            continue
        if not hasattr(cache_field.sharedItems, '_fields'):
            continue

        shared_items = cache_field.sharedItems._fields
        if not shared_items:
            continue

        # Build cache_index -> value mapping (don't modify sharedItems order)
        cache_idx_to_value = {}
        for i, item in enumerate(shared_items):
            if hasattr(item, 'v'):
                cache_idx_to_value[i] = item.v

        # Update all pivot tables using this cache
        for pivot in pivots:
            if field_idx >= len(pivot.pivotFields):
                continue

            pivot_field = pivot.pivotFields[field_idx]
            if pivot_field.items is None:
                continue

            # Separate data items (have a cache index) from subtotal/blank items
            data_items = [item for item in pivot_field.items if item.x is not None]
            other_items = [item for item in pivot_field.items if item.x is None]

            # Sort data items by position in desired_order (unknowns go last)
            def sort_key(item):
                val = cache_idx_to_value.get(item.x, "")
                try:
                    return (0, desired_order.index(val))
                except ValueError:
                    return (1, item.x)  # Unknowns go last, in cache order

            data_items.sort(key=sort_key)
            pivot_field.items = data_items + other_items
            pivot_field.sortType = "manual"
            pivot_field.autoSort = False  # Disable auto-sorting

        # Do NOT set refreshOnLoad - that causes Excel to rebuild item order from cache
        # cache.refreshOnLoad = True


def compact_pivot_tables(workbook, gap_columns=1):
    """Delete extra empty columns between pivot tables to achieve exact spacing.

    Calculates each pivot's width by reading actual unique values from the
    source data table, then deletes extra columns between adjacent pivots.

    Args:
        workbook: openpyxl Workbook
        gap_columns: Number of empty columns to keep between pivot tables (default: 1)
    """
    for worksheet in workbook.worksheets:
        if not hasattr(worksheet, '_pivots') or not worksheet._pivots:
            continue

        pivots = list(worksheet._pivots)
        if len(pivots) < 2:
            continue

        # Sort pivots by their starting column (left to right)
        pivots.sort(key=lambda p: _get_pivot_start_col(p))

        # Track deletions for each pivot gap
        # deletions[i] = number of columns deleted between pivot i and i+1
        deletions = []

        for i in range(len(pivots) - 1):
            curr_pivot = pivots[i]
            next_pivot = pivots[i + 1]

            # Get field value counts from THIS pivot's source data
            field_value_counts = _get_field_value_counts_from_source(workbook, curr_pivot)

            curr_start = _get_pivot_start_col(curr_pivot)
            curr_width = _calculate_pivot_width(curr_pivot, field_value_counts)
            curr_end = curr_start + curr_width - 1  # Last column of current pivot

            next_start = _get_pivot_start_col(next_pivot)

            # Current gap between pivots
            current_gap = next_start - curr_end - 1

            # How many extra columns to delete
            extra_cols = max(0, current_gap - gap_columns)
            deletions.append((curr_end + gap_columns + 1, extra_cols))

        # Delete columns from right to left to preserve indices
        # When deleting right-to-left, positions to the left don't change
        for i in range(len(deletions) - 1, -1, -1):
            delete_start, extra_cols = deletions[i]
            if extra_cols > 0:
                for _ in range(extra_cols):
                    worksheet.delete_cols(delete_start)

        # Update pivot locations - each pivot after the first shifts left
        # by the cumulative deletions that occurred before it
        cumulative_deleted = 0
        for i in range(len(deletions)):
            _, extra_cols = deletions[i]
            cumulative_deleted += extra_cols
            if cumulative_deleted > 0:
                next_pivot = pivots[i + 1]
                old_start = _get_pivot_start_col(next_pivot)
                new_start = old_start - cumulative_deleted
                _set_pivot_start_col(next_pivot, new_start)
                current_app.logger.info(
                    f"Shifted pivot '{next_pivot.name}' from col {old_start} to {new_start}"
                )


def _get_field_value_counts_from_source(workbook, pivot):
    """Read actual unique value counts from the pivot's source data table.

    Returns:
        Dict mapping field names to their unique value count in the data.
    """
    field_value_counts = {}

    if not pivot.cache or not pivot.cache.cacheSource:
        current_app.logger.info("No cache or cacheSource found")
        return field_value_counts

    # Get source worksheet and range from cache
    source = pivot.cache.cacheSource
    if not hasattr(source, 'worksheetSource') or not source.worksheetSource:
        current_app.logger.info("No worksheetSource found")
        return field_value_counts

    ws_source = source.worksheetSource

    # Try different attribute names
    source_sheet_name = getattr(ws_source, 'sheet', None) or getattr(ws_source, 'name', None)
    source_ref = getattr(ws_source, 'ref', None)

    if not source_sheet_name or source_sheet_name not in workbook.sheetnames:
        return field_value_counts

    source_ws = workbook[source_sheet_name]

    # Find the data range - could be a named table or cell range
    min_col = max_col = min_row = max_row = None

    if source_ref and ":" in source_ref:
        # It's a cell range like "A1:L100"
        start_ref, end_ref = source_ref.split(":")
        start_col_letter, start_row = coordinate_from_string(start_ref)
        end_col_letter, end_row = coordinate_from_string(end_ref)
        min_col = column_index_from_string(start_col_letter)
        max_col = column_index_from_string(end_col_letter)
        min_row = int(start_row)
        max_row = int(end_row)
    else:
        # Try to find a table
        table = None
        for t in source_ws.tables.values():
            if source_ref and t.name == source_ref:
                table = t
                break
        if not table:
            table = next(iter(source_ws.tables.values()), None)

        if table and table.ref:
            start_ref, end_ref = table.ref.split(":")
            start_col_letter, start_row = coordinate_from_string(start_ref)
            end_col_letter, end_row = coordinate_from_string(end_ref)
            min_col = column_index_from_string(start_col_letter)
            max_col = column_index_from_string(end_col_letter)
            min_row = int(start_row)
            max_row = int(end_row)

    if min_col is None:
        return field_value_counts

    # Build header -> column index mapping
    headers = {}
    for col in range(min_col, max_col + 1):
        header_value = source_ws.cell(row=min_row, column=col).value
        if header_value:
            headers[header_value] = col

    # For each cache field, count unique values in that column
    if pivot.cache.cacheFields:

        for cache_field in pivot.cache.cacheFields:
            field_name = cache_field.name
            if field_name not in headers:
                continue

            col_idx = headers[field_name]
            unique_values = set()

            for row in range(min_row + 1, max_row + 1):
                cell_value = source_ws.cell(row=row, column=col_idx).value
                if cell_value is not None and cell_value != "":
                    unique_values.add(cell_value)
                else:
                    # Track that there are blank/null values
                    unique_values.add("__BLANK__")

            # Count includes "(blank)" column if there are any null/empty values
            value_count = len(unique_values)
            field_value_counts[field_name] = value_count

    return field_value_counts


def _get_pivot_start_col(pivot):
    """Get the starting column index (1-based) of a pivot table."""
    if pivot.location and pivot.location.ref:
        ref = pivot.location.ref
        # Handle range refs like "E4:G7" - take the start cell
        if ":" in ref:
            ref = ref.split(":")[0]
        col_letter, _ = coordinate_from_string(ref)
        return column_index_from_string(col_letter)
    return 1


def _set_pivot_start_col(pivot, new_col):
    """Set the starting column of a pivot table, preserving range structure."""
    if pivot.location and pivot.location.ref:
        ref = pivot.location.ref
        if ":" in ref:
            # Range ref like "E4:G7" - update both parts
            start_ref, end_ref = ref.split(":")
            start_col_letter, start_row = coordinate_from_string(start_ref)
            end_col_letter, end_row = coordinate_from_string(end_ref)

            # Calculate column offset
            old_start_col = column_index_from_string(start_col_letter)
            old_end_col = column_index_from_string(end_col_letter)
            col_offset = new_col - old_start_col

            new_end_col = old_end_col + col_offset
            pivot.location.ref = (
                f"{get_column_letter(new_col)}{start_row}:"
                f"{get_column_letter(new_end_col)}{end_row}"
            )
        else:
            _, row = coordinate_from_string(ref)
            pivot.location.ref = f"{get_column_letter(new_col)}{row}"


def _calculate_pivot_width(pivot, field_value_counts=None):
    """Calculate the rendered width of a pivot table in columns."""
    if field_value_counts is None:
        field_value_counts = {}

    # Count actual row fields (excluding special values like -2 for "Values")
    actual_row_fields = 0
    if pivot.rowFields:
        actual_row_fields = sum(1 for rf in pivot.rowFields if rf.x is not None and rf.x >= 0)

    # Check if pivot is in compact form (default) or tabular/outline form
    # In compact form, all row fields share one column
    # In tabular/outline form, each row field gets its own column
    is_compact = True
    if pivot.pivotFields and actual_row_fields > 0:
        # Check if any row field is NOT in compact mode
        for rf in pivot.rowFields:
            if rf.x is not None and rf.x >= 0 and rf.x < len(pivot.pivotFields):
                pf = pivot.pivotFields[rf.x]
                # compact controls column layout:
                # - compact=True (default): all row fields share one column
                # - compact=False: each row field gets its own column
                # outline controls subtotal display, NOT column layout
                if pf.compact is False:
                    is_compact = False
                    break

    if is_compact:
        row_field_count = 1  # All row fields share one column
    else:
        row_field_count = max(actual_row_fields, 1)  # Each row field gets its own column

    # Count data fields
    data_field_count = len(pivot.dataFields) if pivot.dataFields else 1

    # Check if there are column fields
    has_col_fields = False
    col_value_count = 1

    if pivot.colFields and pivot.cache:
        for col_field in pivot.colFields:
            if col_field.x is None or col_field.x < 0:
                continue  # Skip "Values" placeholder (-2)
            has_col_fields = True
            field_idx = col_field.x
            if field_idx < len(pivot.cache.cacheFields):
                cache_field = pivot.cache.cacheFields[field_idx]
                field_name = cache_field.name

                # Use actual data count if available
                if field_name in field_value_counts:
                    field_values = field_value_counts[field_name]
                elif hasattr(cache_field, 'sharedItems') and cache_field.sharedItems:
                    if hasattr(cache_field.sharedItems, '_fields'):
                        field_values = len(cache_field.sharedItems._fields)
                    else:
                        field_values = 1
                else:
                    field_values = 1

                col_value_count *= max(field_values, 1)

    # If no column fields, width is just row labels + data fields
    if not has_col_fields:
        total = row_field_count + data_field_count
        return total

    # Calculate data area width (col values × data fields)
    data_width = col_value_count * data_field_count

    # Grand total column(s)
    grand_total_width = 0
    if pivot.colGrandTotals is None or pivot.colGrandTotals:
        grand_total_width = data_field_count

    total = row_field_count + data_width + grand_total_width

    return total
